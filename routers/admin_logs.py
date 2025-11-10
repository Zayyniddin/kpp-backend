from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from datetime import datetime

from database import get_db
from models import entry_record, exit_record, warehouse
from models.user import User
from core.roles import check_role
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/admin/logs", tags=["Admin Logs"])


# -------------------- ВЪЕЗДЫ --------------------
@router.get("/entries")
def get_entries(
    current_user: User = Depends(check_role("admin")),
    db: Session = Depends(get_db),
    warehouse_id: int | None = Query(None, description="Фильтр по складу"),
    start_date: datetime | None = Query(None, description="Начальная дата"),
    end_date: datetime | None = Query(None, description="Конечная дата"),
    skip: int = 0,
    limit: int = 20,
):
    query = db.query(entry_record.EntryRecord).order_by(entry_record.EntryRecord.created_at.desc())

    if warehouse_id:
        query = query.filter(entry_record.EntryRecord.warehouse_id == warehouse_id)
    if start_date:
        query = query.filter(entry_record.EntryRecord.created_at >= start_date)
    if end_date:
        query = query.filter(entry_record.EntryRecord.created_at <= end_date)

    entries = query.offset(skip).limit(limit).all()
    warehouse_map = {w.id: w.name for w in db.query(warehouse.Warehouse).all()}

    return [
        {
            "id": e.id,
            "plate_number": e.plate_number,
            "driver_name": e.driver_name,
            "driver_phone": e.driver_phone,
            "project": e.project,
            "comment": e.comment,
            "created_at": e.created_at,
            "warehouse_name": warehouse_map.get(e.warehouse_id, "—"),
        }
        for e in entries
    ]


# -------------------- ВЫЕЗДЫ --------------------
@router.get("/exits")
def get_exits(
    current_user: User = Depends(check_role("admin")),
    db: Session = Depends(get_db),
    warehouse_id: int | None = Query(None, description="Фильтр по складу"),
    start_date: datetime | None = Query(None, description="Начальная дата"),
    end_date: datetime | None = Query(None, description="Конечная дата"),
    skip: int = 0,
    limit: int = 20,
):
    query = db.query(exit_record.ExitRecord).order_by(exit_record.ExitRecord.exit_time.desc())

    if warehouse_id:
        query = query.filter(exit_record.ExitRecord.warehouse_id == warehouse_id)
    if start_date:
        query = query.filter(exit_record.ExitRecord.exit_time >= start_date)
    if end_date:
        query = query.filter(exit_record.ExitRecord.exit_time <= end_date)

    exits = query.offset(skip).limit(limit).all()
    warehouse_map = {w.id: w.name for w in db.query(warehouse.Warehouse).all()}

    return [
        {
            "id": e.id,
            "pass_number": e.pass_number,
            "places_count": e.places_count,
            "direction": e.direction,
            "project": e.project,
            "comment": e.comment,
            "exit_time": e.exit_time,
            "warehouse_name": warehouse_map.get(e.warehouse_id, "—"),
        }
        for e in exits
    ]


@router.get("/export")
def export_report(
    current_user: User = Depends(check_role("admin")),
    db: Session = Depends(get_db),
    type: str = Query("both", enum=["entries", "exits", "both"], description="Тип отчёта"),
    warehouse_id: int | None = Query(None),
    start_date: datetime | None = Query(None),
    end_date: datetime | None = Query(None),
):
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    has_entries = type in ("entries", "both")
    has_exits = type in ("exits", "both")

    # ---------------- Въезды ----------------
    if has_entries:
        ws1 = wb.active
        ws1.title = "Въезды"
        entries = db.query(entry_record.EntryRecord)
        if warehouse_id:
            entries = entries.filter(entry_record.EntryRecord.warehouse_id == warehouse_id)
        if start_date:
            entries = entries.filter(entry_record.EntryRecord.created_at >= start_date)
        if end_date:
            entries = entries.filter(entry_record.EntryRecord.created_at <= end_date)
        entries = entries.all()

        ws1.append(["ID", "Дата", "Номер", "Водитель", "Телефон", "Проект", "Комментарий", "Склад"])
        for e in entries:
            wh = db.get(warehouse.Warehouse, e.warehouse_id)
            ws1.append([
                e.id,
                e.created_at.strftime("%d.%m.%Y %H:%M"),
                e.plate_number,
                e.driver_name,
                e.driver_phone,
                e.project,
                e.comment,
                wh.name if wh else "—",
            ])
    else:
        wb.remove(wb.active)

    # ---------------- Выезды ----------------
    if has_exits:
        ws2 = wb.create_sheet("Выезды")
        exits = db.query(exit_record.ExitRecord)
        if warehouse_id:
            exits = exits.filter(exit_record.ExitRecord.warehouse_id == warehouse_id)
        if start_date:
            exits = exits.filter(exit_record.ExitRecord.exit_time >= start_date)
        if end_date:
            exits = exits.filter(exit_record.ExitRecord.exit_time <= end_date)
        exits = exits.all()

        ws2.append(["ID", "Дата", "Пропуск", "Мест", "Направление", "Проект", "Комментарий", "Склад"])
        for e in exits:
            wh = db.get(warehouse.Warehouse, e.warehouse_id)
            ws2.append([
                e.id,
                e.exit_time.strftime("%d.%m.%Y %H:%M") if e.exit_time else "—",
                e.pass_number,
                e.places_count,
                e.direction,
                e.project,
                e.comment,
                wh.name if wh else "—",
            ])

    # ---------------- Форматирование ----------------
    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # ---------------- Ответ ----------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"kpp_{type}_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )